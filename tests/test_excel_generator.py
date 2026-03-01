"""Testes para formatacao de KM/local e estrutura do relatorio Excel."""

import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from report.excel_generator import _formatar_km_local, _formatar_trecho_especifico, gerar_relatorio


def test_formatar_trecho_seta_para_entre():
    assert _formatar_trecho_especifico("Resende -> Rio de Janeiro") == "entre Resende e Rio de Janeiro"


def test_formatar_km_local_proximo():
    txt = _formatar_km_local(290.3, "proximo a Itapecerica da Serra")
    assert txt == "KM 290.3 - proximo a Itapecerica da Serra"


def test_formatar_km_local_entre():
    txt = _formatar_km_local(12.7, "Sao Bernardo do Campo -> Riacho Grande")
    assert txt == "KM 12.7 - entre Sao Bernardo do Campo e Riacho Grande"


def test_gera_tres_abas_no_modo_simplificado(tmp_path):
    caminho = gerar_relatorio(
        dados_correlacionados=[],
        pasta_saida=str(tmp_path),
        prefixo="teste_mvp",
        modo_simplificado=True,
    )

    wb = openpyxl.load_workbook(caminho)
    assert wb.sheetnames == ["Monitoramento", "Incidentes", "Resumo"]


def test_gera_tres_abas_no_modo_completo(tmp_path):
    caminho = gerar_relatorio(
        dados_correlacionados=[],
        pasta_saida=str(tmp_path),
        prefixo="teste_full",
        modo_simplificado=False,
    )

    wb = openpyxl.load_workbook(caminho)
    assert wb.sheetnames == ["Monitoramento", "Incidentes", "Resumo"]
