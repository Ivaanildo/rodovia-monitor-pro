"""Excel report generator for monitored routes."""

import os
import unicodedata
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

CORES = {
    "header_bg": "0F4C81",
    "header_font": "FFFFFF",
    "title_bg": "EAF2FA",
    "subtitle_font": "4F5D6B",
    "normal_bg": "D5F5E3",
    "normal_font": "1E8449",
    "moderado_bg": "FEF9E7",
    "moderado_font": "9A7D0A",
    "intenso_bg": "FDEDEC",
    "intenso_font": "B03A2E",
    "interdicao_bg": "F4ECF7",
    "interdicao_font": "6C3483",
    "acidente_bg": "FADBD8",
    "acidente_font": "922B21",
    "bloqueio_parcial_bg": "FDEBD0",
    "bloqueio_parcial_font": "784212",
    "parado_bg": "E5E8E8",
    "parado_font": "1B2631",
    "erro_bg": "E5E7E9",
    "erro_font": "5D6D7E",
    "zebra_bg": "F8FAFC",
    "link_font": "1F5A99",
    "alta_bg": "D4EFDF",
    "alta_font": "145A32",
    "media_bg": "FCF3CF",
    "media_font": "7D6608",
    "baixa_bg": "FADBD8",
    "baixa_font": "922B21",
}

HEADER_FILL = PatternFill("solid", fgColor=CORES["header_bg"])
TITLE_FILL = PatternFill("solid", fgColor=CORES["title_bg"])
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=CORES["header_font"])
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=CORES["header_bg"])
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color=CORES["subtitle_font"])
DEFAULT_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(name="Calibri", size=10, color=CORES["link_font"], underline="single")
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ZEBRA_FILL = PatternFill("solid", fgColor=CORES["zebra_bg"])


def _norm(txt):
    base = unicodedata.normalize("NFKD", str(txt or ""))
    sem_acento = "".join(ch for ch in base if not unicodedata.combining(ch))
    return sem_acento.strip().lower()


# Lookup table unificada: (categoria, valor_normalizado) -> (bg_key, font_key)
_STYLE_MAP = {
    "status": {
        "normal": ("normal_bg", "normal_font"),
        "moderado": ("moderado_bg", "moderado_font"),
        "intenso": ("intenso_bg", "intenso_font"),
        "parado": ("parado_bg", "parado_font"),
    },
    "ocorrencia": {
        "colisao": ("acidente_bg", "acidente_font"),
        "acidente": ("acidente_bg", "acidente_font"),
        "interdicao": ("interdicao_bg", "interdicao_font"),
        "bloqueio parcial": ("bloqueio_parcial_bg", "bloqueio_parcial_font"),
        "obras na pista": ("moderado_bg", "moderado_font"),
        "engarrafamento": ("intenso_bg", "intenso_font"),
    },
    "confianca": {
        "alta": ("alta_bg", "alta_font"),
        "media": ("media_bg", "media_font"),
        "média": ("media_bg", "media_font"),
        "baixa": ("baixa_bg", "baixa_font"),
    },
    "confianca_loc": {
        "alta": ("alta_bg", "alta_font"),
        "media": ("media_bg", "media_font"),
        "baixa": ("baixa_bg", "baixa_font"),
    },
}


def _get_style(categoria, valor):
    """Retorna (fill, font) para uma categoria e valor. Fallback por categoria."""
    key = _norm(valor)
    mapping = _STYLE_MAP.get(categoria, {})
    colors = mapping.get(key)
    if colors:
        bg_key, fg_key = colors
        return (
            PatternFill("solid", fgColor=CORES[bg_key]),
            Font(name="Calibri", size=10, bold=True, color=CORES[fg_key]),
        )
    # Fallback: status sempre tem cor, outros retornam None
    if categoria == "status":
        return (
            PatternFill("solid", fgColor=CORES["erro_bg"]),
            Font(name="Calibri", size=10, bold=True, color=CORES["erro_font"]),
        )
    return None, None


def _estilizar_header(ws, row_idx, headers):
    for col_idx, (header, width) in enumerate(headers, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _aplicar_linha_base(ws, row_idx, total_cols, zebra=False):
    for col_idx in range(1, total_cols + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.font = DEFAULT_FONT
        c.border = THIN_BORDER
        c.alignment = CENTER
        if zebra:
            c.fill = ZEBRA_FILL


def _texto_curto(texto, limite=170):
    valor = " ".join(str(texto or "").split())
    if len(valor) <= limite:
        return valor
    return valor[: limite - 3].rstrip() + "..."


def _formatar_trecho_especifico(trecho):
    txt = " ".join(str(trecho or "").split())
    if not txt:
        return ""

    low = txt.lower()
    if "->" in txt:
        partes = [p.strip() for p in txt.split("->", 1)]
        if len(partes) == 2 and partes[0] and partes[1]:
            return f"entre {partes[0]} e {partes[1]}"
    if low.startswith("proximo a "):
        return txt
    if low.startswith("entre "):
        return txt
    return txt


def _formatar_km_local(km, trecho_especifico, confianca_loc=None):
    trecho_fmt = _formatar_trecho_especifico(trecho_especifico)
    km_txt = f"KM {km}" if km is not None else ""
    conf_txt = f" ({int(confianca_loc * 100)}%)" if confianca_loc and confianca_loc > 0 else ""

    if km_txt and trecho_fmt:
        return f"{km_txt} - {trecho_fmt}{conf_txt}"
    if km_txt:
        return f"{km_txt}{conf_txt}"
    return trecho_fmt


def _nivel_confianca_loc(confianca_loc):
    """Classifica confianca_localizacao (0.0-1.0) em nivel textual."""
    if confianca_loc is None or confianca_loc <= 0:
        return None
    if confianca_loc >= 0.7:
        return "alta"
    if confianca_loc >= 0.4:
        return "media"
    return "baixa"


def gerar_relatorio(
    dados_correlacionados,
    pasta_saida="./relatorios",
    prefixo="rodoviamonitor_pro",
    modo_simplificado=False,
    resumo_coleta=None,
):
    os.makedirs(pasta_saida, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho = os.path.join(pasta_saida, f"{prefixo}_{ts}.xlsx")

    wb = Workbook()
    if modo_simplificado:
        _gerar_relatorio_simplificado(wb, dados_correlacionados, resumo_coleta=resumo_coleta)
    else:
        _gerar_aba_monitoramento(wb.active, dados_correlacionados, resumo_coleta=resumo_coleta)
    _gerar_aba_incidentes(wb, dados_correlacionados)
    _gerar_aba_resumo(wb, dados_correlacionados)

    wb.save(caminho)
    return caminho


def _gerar_aba_monitoramento(ws, dados, resumo_coleta=None):
    ws.title = "Monitoramento"

    headers = [
        ("#", 5),
        ("Rodovia", 16),
        ("Trecho", 34),
        ("Tipo", 10),
        ("Concessionaria", 26),
        ("Sentido", 20),
        ("KM", 10),
        ("Trecho especifico", 32),
        ("Waze", 14),
        ("Google Maps", 16),
        ("Seção consultada (ponto a ponto)", 52),
        ("Status", 12),
        ("Ocorrencia", 28),
        ("Descricao / Observacoes", 58),
        ("Duracao normal", 14),
        ("Duracao atual", 14),
        ("Atraso (min)", 12),
        ("Jam Factor", 11),
        ("Fontes", 24),
        ("Confianca", 12),
        ("Validacao", 34),
        ("Atualizado em", 20),
    ]
    total_cols = len(headers)
    last_col = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = f"RodoviaMonitor Pro - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    s = ws["A2"]
    s.value = "Dados correlacionados das fontes configuradas"
    s.font = SUBTITLE_FONT
    s.alignment = CENTER

    header_row = 4

    if resumo_coleta and resumo_coleta.get("cobertura_pct", 100) < 50:
        ws.merge_cells(f"A3:{last_col}3")
        aviso = ws["A3"]
        aviso.value = (
            f"ATENCAO: Cobertura de dados {resumo_coleta['cobertura_pct']}% - "
            f"{resumo_coleta['sem_dados']}/{resumo_coleta['total']} trechos sem dados. "
            "Verifique as API keys."
        )
        aviso.font = Font(name="Calibri", size=10, bold=True, color=CORES["acidente_font"])
        aviso.fill = PatternFill("solid", fgColor=CORES["acidente_bg"])
        aviso.alignment = CENTER
        header_row = 5

    _estilizar_header(ws, header_row, headers)

    for idx, d in enumerate(dados, 1):
        row = header_row + idx
        _aplicar_linha_base(ws, row, total_cols, zebra=(idx % 2 == 0))

        km = d.get("km_ocorrencia")
        km_display = f"KM {km}" if km is not None else ""
        values = [
            idx,
            d.get("rodovia", ""),
            d.get("trecho", ""),
            d.get("tipo", "federal"),
            d.get("concessionaria", ""),
            d.get("sentido", ""),
            km_display,
            _formatar_trecho_especifico(d.get("trecho_especifico", "")),
            d.get("link_waze", ""),
            d.get("link_gmaps", ""),
            d.get("trecho_consultado_descricao", ""),
            d.get("status", "Sem dados"),
            d.get("ocorrencia", ""),
            " ".join(str(d.get("descricao", "") or "").split()),
            d.get("duracao_normal_min", "") or "",
            d.get("duracao_transito_min", "") or "",
            d.get("atraso_min", "") or "",
            d.get("jam_factor", "") or "",
            ", ".join(d.get("fontes_utilizadas", [])),
            d.get("confianca", ""),
            d.get("acao_recomendada", ""),
            d.get("consultado_em", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            c = ws.cell(row=row, column=col_idx, value=value)
            if col_idx in (3, 8, 11, 13, 14, 21):
                c.alignment = LEFT
            if col_idx == 9 and values[8]:
                c.value = "Abrir Waze"
                c.hyperlink = values[8]
                c.font = LINK_FONT
            if col_idx == 10 and values[9]:
                c.value = "Abrir Maps"
                c.hyperlink = values[9]
                c.font = LINK_FONT

        status_fill, status_font = _get_style("status", values[11])
        ws.cell(row=row, column=12).fill = status_fill
        ws.cell(row=row, column=12).font = status_font

        occ_principal = d.get("ocorrencia_principal", values[12])
        occ_fill, occ_font = _get_style("ocorrencia", occ_principal)
        if occ_fill and occ_font:
            ws.cell(row=row, column=13).fill = occ_fill
            ws.cell(row=row, column=13).font = occ_font

        conf_fill, conf_font = _get_style("confianca", values[19])
        if conf_fill and conf_font:
            ws.cell(row=row, column=20).fill = conf_fill
            ws.cell(row=row, column=20).font = conf_font

        descricao_len = len(str(values[13] or ""))
        occ_len = len(str(values[12] or ""))
        max_content_len = max(descricao_len, occ_len)
        ws.row_dimensions[row].height = min(54, 22 + (max_content_len // 85) * 8)

    legenda_start = header_row + len(dados) + 3
    ws.cell(row=legenda_start, column=1, value="Legenda").font = Font(name="Calibri", size=10, bold=True)

    legendas = [
        ("Status Normal", CORES["normal_bg"], CORES["normal_font"]),
        ("Status Moderado", CORES["moderado_bg"], CORES["moderado_font"]),
        ("Status Intenso", CORES["intenso_bg"], CORES["intenso_font"]),
        ("Status Parado", CORES["parado_bg"], CORES["parado_font"]),
        ("Ocorrencia Colisao", CORES["acidente_bg"], CORES["acidente_font"]),
        ("Ocorrencia Interdicao", CORES["interdicao_bg"], CORES["interdicao_font"]),
        ("Ocorrencia Bloqueio Parcial", CORES["bloqueio_parcial_bg"], CORES["bloqueio_parcial_font"]),
    ]
    for offset, (label, bg, fg) in enumerate(legendas, 1):
        c = ws.cell(row=legenda_start + offset, column=1, value=label)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", size=9, bold=True, color=fg)
        c.border = THIN_BORDER

    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(dados)}"
    ws.freeze_panes = f"A{header_row + 1}"


def _gerar_aba_incidentes(wb, dados):
    ws = wb.create_sheet("Incidentes")
    headers = [
        ("#", 5),
        ("Trecho", 30),
        ("Fonte", 14),
        ("Categoria", 18),
        ("Severidade", 12),
        ("Rodovia", 16),
        ("Sentido", 16),
        ("KM", 11),
        ("Local", 30),
        ("Descricao", 56),
        ("Horario", 20),
    ]
    _estilizar_header(ws, 1, headers)

    linhas = []
    for d in dados:
        for inc in d.get("incidentes_detalhados", []):
            km_inc = inc.get("km_estimado")
            linhas.append([
                d.get("trecho", ""),
                inc.get("fonte", ""),
                inc.get("categoria", ""),
                inc.get("severidade", ""),
                inc.get("rodovia_afetada", d.get("rodovia", "")),
                inc.get("sentido", ""),
                f"KM {km_inc}" if km_inc is not None else "",
                inc.get("trecho_especifico", "") or inc.get("localizacao_precisa", ""),
                inc.get("descricao", ""),
                inc.get("consultado_em", ""),
            ])

    if not linhas:
        ws.cell(row=2, column=1, value="Nenhum incidente encontrado nesta consulta.").font = Font(
            name="Calibri", size=10, italic=True
        )
    else:
        for idx, linha in enumerate(linhas, 1):
            row = idx + 1
            _aplicar_linha_base(ws, row, len(headers), zebra=(idx % 2 == 0))
            ws.cell(row=row, column=1, value=idx)
            for col_idx, value in enumerate(linha, 2):
                c = ws.cell(row=row, column=col_idx, value=value)
                c.alignment = LEFT if col_idx in (2, 9, 10) else CENTER

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, len(linhas) + 1)}"
    ws.freeze_panes = "A2"


def _gerar_aba_resumo(wb, dados):
    ws = wb.create_sheet("Resumo")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28

    ws["A1"] = "Resumo Executivo"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Data/Hora"
    ws["A3"].font = HEADER_FONT
    ws["A3"].fill = HEADER_FILL
    ws["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    ws["A4"] = "Trechos monitorados"
    ws["A4"].font = HEADER_FONT
    ws["A4"].fill = HEADER_FILL
    ws["B4"] = len(dados)

    status_count = {}
    ocorrencia_count = {}
    confianca_count = {}
    for d in dados:
        status = d.get("status", "Sem dados")
        ocorrencia_raw = d.get("ocorrencia", "") or ""
        confianca = d.get("confianca", "Baixa")
        status_count[status] = status_count.get(status, 0) + 1
        if ocorrencia_raw:
            for occ in ocorrencia_raw.split(";"):
                occ = occ.strip()
                if occ:
                    ocorrencia_count[occ] = ocorrencia_count.get(occ, 0) + 1
        else:
            ocorrencia_count["Sem ocorrencia"] = (
                ocorrencia_count.get("Sem ocorrencia", 0) + 1
            )
        confianca_count[confianca] = confianca_count.get(confianca, 0) + 1

    row = 6
    ws.cell(row=row, column=1, value="Status").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for status, qty in sorted(status_count.items(), key=lambda item: (-item[1], item[0])):
        ws.cell(row=row, column=1, value=status).border = THIN_BORDER
        ws.cell(row=row, column=2, value=qty).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Ocorrencias").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for ocorrencia, qty in sorted(ocorrencia_count.items(), key=lambda item: (-item[1], item[0])):
        ws.cell(row=row, column=1, value=ocorrencia).border = THIN_BORDER
        ws.cell(row=row, column=2, value=qty).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Confianca").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for confianca, qty in sorted(confianca_count.items(), key=lambda item: (-item[1], item[0])):
        ws.cell(row=row, column=1, value=confianca).border = THIN_BORDER
        ws.cell(row=row, column=2, value=qty).border = THIN_BORDER


def _gerar_relatorio_simplificado(wb, dados, resumo_coleta=None):
    ws = wb.active
    ws.title = "Monitoramento"

    headers = [
        ("Rodovia", 16),
        ("Trecho", 34),
        ("Sentido", 20),
        ("Status", 12),
        ("Ocorrencia", 28),
        ("Observacoes", 58),
        ("Seção consultada (ponto a ponto)", 52),
        ("Google Maps", 16),
        ("Atualizado em", 20),
    ]

    total_cols = len(headers)
    last_col = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"Monitoramento - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = CENTER

    header_row = 3

    if resumo_coleta and resumo_coleta.get("cobertura_pct", 100) < 50:
        ws.merge_cells(f"A2:{last_col}2")
        aviso = ws["A2"]
        aviso.value = (
            f"ATENCAO: Cobertura de dados {resumo_coleta['cobertura_pct']}% - "
            f"{resumo_coleta['sem_dados']}/{resumo_coleta['total']} trechos sem dados. "
            "Verifique as API keys."
        )
        aviso.font = Font(name="Calibri", size=10, bold=True, color=CORES["acidente_font"])
        aviso.fill = PatternFill("solid", fgColor=CORES["acidente_bg"])
        aviso.alignment = CENTER
        header_row = 4

    _estilizar_header(ws, header_row, headers)

    for idx, d in enumerate(dados, 1):
        row = header_row + idx
        _aplicar_linha_base(ws, row, total_cols, zebra=(idx % 2 == 0))

        values = [
            d.get("rodovia", ""),
            d.get("trecho", ""),
            d.get("sentido", ""),
            d.get("status", "Sem dados"),
            d.get("ocorrencia", ""),
            " ".join(str(d.get("descricao", "") or "").split()),
            d.get("trecho_consultado_descricao", ""),
            d.get("link_gmaps", ""),
            d.get("consultado_em", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            c = ws.cell(row=row, column=col_idx, value=value)
            c.alignment = LEFT if col_idx in (2, 6, 7) else CENTER
            if col_idx == 8 and values[7]:
                c.value = "Abrir Maps"
                c.hyperlink = values[7]
                c.font = LINK_FONT

        status_fill, status_font = _get_style("status", values[3])
        ws.cell(row=row, column=4).fill = status_fill
        ws.cell(row=row, column=4).font = status_font

        occ_principal = d.get("ocorrencia_principal", values[4])
        occ_fill, occ_font = _get_style("ocorrencia", occ_principal)
        if occ_fill and occ_font:
            ws.cell(row=row, column=5).fill = occ_fill
            ws.cell(row=row, column=5).font = occ_font

        obs_len = len(str(values[5] or ""))
        occ_len = len(str(values[4] or ""))
        max_len = max(obs_len, occ_len)
        ws.row_dimensions[row].height = min(90, 22 + (max_len // 60) * 8)

    max_row = max(header_row + 1, header_row + len(dados))
    
    range_status = f"D{header_row + 1}:D{max_row}"
    dv_status = DataValidation(type="list", formula1='"Normal,Moderado,Intenso,Parado"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(range_status)

    opcoes_status = {
        "Normal": ("normal_bg", "normal_font"),
        "Moderado": ("moderado_bg", "moderado_font"),
        "Intenso": ("intenso_bg", "intenso_font"),
        "Parado": ("intenso_bg", "intenso_font"),
    }

    for op_texto, (bg_key, font_key) in opcoes_status.items():
        rule = CellIsRule(operator='equal', formula=[f'"{op_texto}"'], stopIfTrue=True,
                          fill=PatternFill("solid", fgColor=CORES[bg_key]),
                          font=Font(name="Calibri", size=10, bold=True, color=CORES[font_key]))
        ws.conditional_formatting.add(range_status, rule)

    range_ocorrencia = f"E{header_row + 1}:E{max_row}"
    dv = DataValidation(type="list", formula1='"Colisão,Obras na Pista,Engarrafamento,Interdição"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(range_ocorrencia)

    opcoes = {
        "Colisão": ("acidente_bg", "acidente_font"),
        "Obras na Pista": ("moderado_bg", "moderado_font"),
        "Engarrafamento": ("intenso_bg", "intenso_font"),
        "Interdição": ("interdicao_bg", "interdicao_font"),
    }

    for op_texto, (bg_key, font_key) in opcoes.items():
        rule = CellIsRule(operator='equal', formula=[f'"{op_texto}"'], stopIfTrue=True,
                          fill=PatternFill("solid", fgColor=CORES[bg_key]),
                          font=Font(name="Calibri", size=10, bold=True, color=CORES[font_key]))
        ws.conditional_formatting.add(range_ocorrencia, rule)

    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(dados)}"
    ws.freeze_panes = f"A{header_row + 1}"
