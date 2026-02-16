#!/usr/bin/env python3
"""Validador factual e estrutural das rotas do RodoviaMonitor."""

from __future__ import annotations

import argparse
import csv
import json
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List
from urllib.parse import urlparse

import openpyxl
import yaml

ROUTES_ESPERADAS = 28
LIMITE_GAP_INTERMUNICIPAL_KM = 60.0
LIMITE_GAP_METROPOLITANO_KM = 20.0
_HINTS_METRO = (
    "urbano",
    "marginal",
    "avenida brasil",
    "recife",
    "olinda",
    "paulista",
    "deodoro",
    "jaguare",
    "jaguaré",
    "sao paulo",
    "são paulo",
    "taboao",
    "taboão",
    "embu",
    "itapecerica",
    "diadema",
    "sao bernardo",
    "são bernardo",
)
_CORREDOR_KEYS = {"corredor", "corredor principal", "corredor_principal"}


@dataclass
class RouteReport:
    rota: str
    status: str = "Aprovado"
    erros: List[str] = field(default_factory=list)
    alertas: List[str] = field(default_factory=list)

    def finalize(self) -> None:
        if self.erros:
            self.status = "Reprovado"
        elif self.alertas:
            self.status = "Aprovado com ressalva documentada"
        else:
            self.status = "Aprovado"


def _norm(texto: str) -> str:
    txt = unicodedata.normalize("NFKD", str(texto or ""))
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    return " ".join(txt.lower().split())


def _limite_gap_rota(rota: dict) -> float:
    if rota.get("limite_gap_km") is not None:
        return float(rota["limite_gap_km"])

    chave = _norm(rota.get("nome", ""))
    if any(h in chave for h in _HINTS_METRO):
        return LIMITE_GAP_METROPOLITANO_KM
    return LIMITE_GAP_INTERMUNICIPAL_KM


def _carregar_rotas(path_rotas: Path) -> List[dict]:
    data = yaml.safe_load(path_rotas.read_text(encoding="utf-8")) or {}
    rotas = data.get("rotas", [])
    if not isinstance(rotas, list):
        raise ValueError("Campo 'rotas' invalido no YAML.")
    return rotas


def _carregar_evidencias(path_csv: Path) -> Dict[str, Dict[str, List[dict]]]:
    evidencias: Dict[str, Dict[str, List[dict]]] = {}
    if not path_csv.exists():
        return evidencias

    with path_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rota_key = _norm(row.get("rota", ""))
            ponto_key = _norm(row.get("ponto", ""))
            if not rota_key:
                continue
            evidencias.setdefault(rota_key, {}).setdefault(ponto_key, []).append(row)
    return evidencias


def _to_float(value, default=None):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _carregar_linhas_excel(path_excel: Path) -> List[dict]:
    wb = openpyxl.load_workbook(path_excel)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    header_map = {}
    for row_idx in range(1, ws.max_row + 1):
        row_vals = [str(ws.cell(row_idx, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if "Trecho" in row_vals and ("Rodovia" in row_vals or "#" in row_vals):
            header_row = row_idx
            header_map = {v: i + 1 for i, v in enumerate(row_vals) if v}
            break

    if not header_row:
        raise ValueError("Nao foi possivel localizar cabecalho no Excel.")

    idx_rodovia = header_map.get("Rodovia")
    idx_trecho = header_map.get("Trecho")
    idx_sentido = header_map.get("Sentido")
    idx_gmaps = header_map.get("Google Maps")
    if not (idx_rodovia and idx_trecho and idx_sentido and idx_gmaps):
        raise ValueError("Cabecalho do Excel incompleto para validacao.")

    linhas = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        rodovia = ws.cell(row_idx, idx_rodovia).value
        trecho = ws.cell(row_idx, idx_trecho).value
        sentido = ws.cell(row_idx, idx_sentido).value
        cell_maps = ws.cell(row_idx, idx_gmaps)
        link_maps = cell_maps.hyperlink.target if cell_maps.hyperlink else ""

        if not (rodovia or trecho or sentido):
            continue

        linhas.append(
            {
                "row": row_idx,
                "rodovia": str(rodovia or "").strip(),
                "trecho": str(trecho or "").strip(),
                "sentido": str(sentido or "").strip(),
                "link_gmaps": str(link_maps or "").strip(),
            }
        )
    return linhas


def _validar_link_gmaps(link: str) -> bool:
    if not link:
        return False
    parsed = urlparse(link)
    if "google.com" not in parsed.netloc:
        return False
    return parsed.path.startswith("/maps/dir")


def _validar_estrutura(rotas: List[dict], reports: Dict[str, RouteReport], globais: List[str]) -> None:
    if len(rotas) != ROUTES_ESPERADAS:
        globais.append(
            f"Quantidade de rotas divergente: esperado {ROUTES_ESPERADAS}, encontrado {len(rotas)}."
        )

    campos_obrigatorios = ("nome", "origem", "destino", "rodovia", "sentido")
    for rota in rotas:
        rep = reports[rota["nome"]]
        for campo in campos_obrigatorios:
            if not rota.get(campo):
                rep.erros.append(f"Campo obrigatorio ausente: {campo}.")

        segmentos = rota.get("segmentos", [])
        if not segmentos:
            rep.erros.append("Rota sem segmentos.")
            continue

        limite_gap = _limite_gap_rota(rota)
        for idx_seg, seg in enumerate(segmentos, 1):
            pontos = seg.get("pontos_referencia", [])
            if len(pontos) < 2:
                rep.erros.append(f"Segmento {idx_seg} com menos de 2 pontos.")
                continue

            prev_km = None
            for p_idx, p in enumerate(pontos, 1):
                for campo in ("km", "lat", "lng", "local"):
                    if p.get(campo) in (None, ""):
                        rep.erros.append(f"Segmento {idx_seg}, ponto {p_idx} sem campo {campo}.")

                km = _to_float(p.get("km"))
                if km is None:
                    rep.erros.append(f"Segmento {idx_seg}, ponto {p_idx} com KM invalido.")
                    continue

                if prev_km is not None and km < prev_km:
                    rep.erros.append(
                        f"Segmento {idx_seg} com KM nao monotono: {km} apos {prev_km}."
                    )
                if prev_km is not None:
                    gap = km - prev_km
                    if gap > limite_gap:
                        rep.alertas.append(
                            f"Gap alto no segmento {idx_seg}: {gap:.1f} km (limite {limite_gap:.1f} km)."
                        )
                prev_km = km


def _validar_excel(rotas: List[dict], linhas_excel: List[dict], globais: List[str]) -> None:
    if len(linhas_excel) != len(rotas):
        globais.append(
            f"Excel com quantidade de linhas divergente: {len(linhas_excel)} vs {len(rotas)} rotas."
        )

    by_nome = {_norm(r["nome"]): r for r in rotas}
    for linha in linhas_excel:
        rota = by_nome.get(_norm(linha["trecho"]))
        if not rota:
            globais.append(f"Linha {linha['row']} do Excel nao mapeia rota: {linha['trecho']}")
            continue

        if _norm(linha["rodovia"]) != _norm(rota.get("rodovia", "")):
            globais.append(
                f"Linha {linha['row']} rodovia divergente: {linha['rodovia']} vs {rota.get('rodovia')}."
            )
        if _norm(linha["sentido"]) != _norm(rota.get("sentido", "")):
            globais.append(
                f"Linha {linha['row']} sentido divergente: {linha['sentido']} vs {rota.get('sentido')}."
            )
        if not _validar_link_gmaps(linha["link_gmaps"]):
            globais.append(
                f"Linha {linha['row']} link Google Maps fora do padrao canonico: {linha['link_gmaps']}"
            )


def _km_em_faixa(km: float, evidencias: List[dict]) -> bool:
    for ev in evidencias:
        km_min = _to_float(ev.get("km_faixa_min"), default=float("-inf"))
        km_max = _to_float(ev.get("km_faixa_max"), default=float("inf"))
        if km_min <= km <= km_max:
            return True
    return False


def _validar_fatos(rotas: List[dict], evidencias: Dict[str, Dict[str, List[dict]]], reports: Dict[str, RouteReport]):
    for rota in rotas:
        rep = reports[rota["nome"]]
        rota_key = _norm(rota["nome"])
        evid_rota = evidencias.get(rota_key, {})
        if not evid_rota:
            rep.alertas.append("Sem evidencia factual cadastrada para a rota.")
            continue

        evid_corredor = []
        for key in _CORREDOR_KEYS:
            evid_corredor.extend(evid_rota.get(key, []))

        for seg in rota.get("segmentos", []):
            for p in seg.get("pontos_referencia", []):
                ponto_key = _norm(p.get("local", ""))
                km = _to_float(p.get("km"))
                if km is None:
                    continue

                evid_ponto = evid_rota.get(ponto_key, [])
                if evid_ponto:
                    if not _km_em_faixa(km, evid_ponto):
                        rep.erros.append(
                            f"Ponto '{p.get('local')}' com KM {km} fora da faixa de evidencia."
                        )
                    continue

                if evid_corredor:
                    if not _km_em_faixa(km, evid_corredor):
                        rep.erros.append(
                            f"Ponto '{p.get('local')}' com KM {km} fora da faixa do corredor."
                        )
                    else:
                        rep.alertas.append(
                            f"Ponto '{p.get('local')}' sem evidencia local; validado apenas por faixa de corredor."
                        )
                else:
                    rep.alertas.append(
                        f"Ponto '{p.get('local')}' sem evidencia local nem de corredor."
                    )


def _salvar_json(path_saida: Path, payload: dict) -> None:
    path_saida.parent.mkdir(parents=True, exist_ok=True)
    path_saida.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _salvar_md(path_saida: Path, payload: dict) -> None:
    path_saida.parent.mkdir(parents=True, exist_ok=True)
    linhas = []
    linhas.append("# Auditoria factual de rotas")
    linhas.append("")
    linhas.append(f"- Gerado em: {payload['gerado_em']}")
    linhas.append(f"- Rotas avaliadas: {payload['resumo']['rotas_total']}")
    linhas.append(f"- Aprovadas: {payload['resumo']['aprovadas']}")
    linhas.append(f"- Aprovadas com ressalva: {payload['resumo']['ressalvas']}")
    linhas.append(f"- Reprovadas: {payload['resumo']['reprovadas']}")
    linhas.append("")

    if payload["globais"]:
        linhas.append("## Achados globais")
        for item in payload["globais"]:
            linhas.append(f"- {item}")
        linhas.append("")

    linhas.append("## Status por rota")
    linhas.append("| Rota | Status | Erros | Alertas |")
    linhas.append("|---|---|---:|---:|")
    for rota in payload["rotas"]:
        linhas.append(
            f"| {rota['rota']} | {rota['status']} | {len(rota['erros'])} | {len(rota['alertas'])} |"
        )
    linhas.append("")

    linhas.append("## Detalhes")
    for rota in payload["rotas"]:
        if not rota["erros"] and not rota["alertas"]:
            continue
        linhas.append(f"### {rota['rota']}")
        for err in rota["erros"]:
            linhas.append(f"- ERRO: {err}")
        for warn in rota["alertas"]:
            linhas.append(f"- ALERTA: {warn}")
        linhas.append("")

    path_saida.write_text("\n".join(linhas).strip() + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Valida fatos e estrutura das rotas monitoradas.")
    parser.add_argument("--rotas", required=True, help="Arquivo YAML de rotas.")
    parser.add_argument("--excel", required=True, help="Arquivo Excel base para confronto.")
    parser.add_argument("--saida-json", required=True, help="Arquivo JSON de saida.")
    parser.add_argument("--saida-md", required=True, help="Arquivo Markdown de saida.")
    parser.add_argument(
        "--fontes-csv",
        default="docs/factcheck/fontes_rotas_2026-02-16.csv",
        help="Arquivo CSV com evidencias factuais.",
    )
    args = parser.parse_args()

    path_rotas = Path(args.rotas)
    path_excel = Path(args.excel)
    path_json = Path(args.saida_json)
    path_md = Path(args.saida_md)
    path_fontes = Path(args.fontes_csv)

    rotas = _carregar_rotas(path_rotas)
    evidencias = _carregar_evidencias(path_fontes)
    linhas_excel = _carregar_linhas_excel(path_excel)

    globais: List[str] = []
    reports = {r["nome"]: RouteReport(rota=r["nome"]) for r in rotas}

    _validar_estrutura(rotas, reports, globais)
    _validar_excel(rotas, linhas_excel, globais)
    _validar_fatos(rotas, evidencias, reports)

    for rep in reports.values():
        rep.finalize()

    aprovadas = sum(1 for r in reports.values() if r.status == "Aprovado")
    ressalvas = sum(1 for r in reports.values() if r.status == "Aprovado com ressalva documentada")
    reprovadas = sum(1 for r in reports.values() if r.status == "Reprovado")

    payload = {
        "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "arquivos": {
            "rotas": str(path_rotas),
            "excel": str(path_excel),
            "fontes_csv": str(path_fontes),
        },
        "globais": globais,
        "resumo": {
            "rotas_total": len(rotas),
            "aprovadas": aprovadas,
            "ressalvas": ressalvas,
            "reprovadas": reprovadas,
        },
        "rotas": [
            {
                "rota": rep.rota,
                "status": rep.status,
                "erros": rep.erros,
                "alertas": rep.alertas,
            }
            for rep in reports.values()
        ],
    }

    _salvar_json(path_json, payload)
    _salvar_md(path_md, payload)

    return 2 if reprovadas > 0 else 0


if __name__ == "__main__":
    raise SystemExit(main())
